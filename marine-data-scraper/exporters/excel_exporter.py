"""
Export marine company data to a formatted Excel workbook.
Sheets: Summary, All Companies, Contacts & Emails, By Country, By Type.
"""
import logging
from pathlib import Path
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func

from database.models import Company, Contact, Email, VesselType
from database.db import get_stats

logger = logging.getLogger(__name__)


def export_excel(session: Session, path: Path, companies: list[Company] = None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return None

    if companies is None:
        companies = session.query(Company).filter_by(is_active=True).all()

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------
    # Styles
    # ---------------------------------------------------------------
    HEADER_FILL  = PatternFill("solid", fgColor="003366")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
    ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")
    TITLE_FONT   = Font(size=14, bold=True, color="003366")
    CENTER       = Alignment(horizontal="center", vertical="center")
    WRAP         = Alignment(wrap_text=True, vertical="top")
    THIN_BORDER  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header_row(ws, row, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # ---------------------------------------------------------------
    # Sheet 1: Summary
    # ---------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    stats = get_stats(session)

    ws["A1"] = "Marine Company Database – Summary"
    ws["A1"].font = Font(size=16, bold=True, color="003366")
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = CENTER

    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Total Companies";  ws["B4"] = stats["total_companies"]
    ws["A5"] = "Total Contacts";   ws["B5"] = stats["total_contacts"]
    ws["A6"] = "Total Emails";     ws["B6"] = stats["total_emails"]
    ws["A7"] = "Total Vessels";    ws["B7"] = stats["total_vessels"]

    row = 9
    ws.cell(row=row, column=1, value="Company Type"); ws.cell(row=row, column=2, value="Count")
    style_header_row(ws, row, 2)
    for ctype, cnt in sorted(stats["companies_by_type"].items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=ctype or "Unknown")
        ws.cell(row=row, column=2, value=cnt)

    row += 2
    ws.cell(row=row, column=1, value="Top Countries"); ws.cell(row=row, column=2, value="Count")
    style_header_row(ws, row, 2)
    for country, cnt in list(stats["companies_by_country"].items())[:15]:
        row += 1
        ws.cell(row=row, column=1, value=country or "Unknown")
        ws.cell(row=row, column=2, value=cnt)

    autofit(ws)

    # ---------------------------------------------------------------
    # Sheet 2: All Companies
    # ---------------------------------------------------------------
    ws2 = wb.create_sheet("All Companies")
    headers = ["ID", "Company Name", "Type", "Country", "City", "Port",
               "Website", "Fleet Size", "Vessel Types", "Email Count",
               "Description", "Source", "Date Scraped"]
    ws2.append(headers)
    style_header_row(ws2, 1, len(headers))
    ws2.freeze_panes = "A2"

    for i, c in enumerate(companies, start=2):
        fill = ALT_FILL if i % 2 == 0 else None
        vtypes = "; ".join(vt.name for vt in c.vessel_types)
        row_data = [
            c.id, c.name, c.company_type or "", c.country or "",
            c.city or "", c.port or "", c.website or "",
            c.fleet_size or "", vtypes, len(c.emails),
            (c.description or "")[:200],
            c.source_name or "",
            c.date_scraped.strftime("%Y-%m-%d") if c.date_scraped else "",
        ]
        ws2.append(row_data)
        if fill:
            for col in range(1, len(headers) + 1):
                ws2.cell(row=i, column=col).fill = fill

    autofit(ws2)

    # ---------------------------------------------------------------
    # Sheet 3: Contacts & Emails (sales-ready)
    # ---------------------------------------------------------------
    ws3 = wb.create_sheet("Contacts & Emails")
    h3 = ["Company Name", "Type", "Country", "City", "Website",
          "Vessel Types", "Contact Name", "Title", "Email", "Email Type",
          "Source"]
    ws3.append(h3)
    style_header_row(ws3, 1, len(h3))
    ws3.freeze_panes = "A2"

    row_idx = 2
    for c in companies:
        vtypes  = "; ".join(vt.name for vt in c.vessel_types)
        written_emails: set[str] = set()

        for co in c.contacts:
            fill = ALT_FILL if row_idx % 2 == 0 else None
            row_data = [
                c.name, c.company_type or "", c.country or "", c.city or "",
                c.website or "", vtypes,
                co.name or "", co.title or "", co.email or "", "contact",
                c.source_name or "",
            ]
            ws3.append(row_data)
            if fill:
                for col in range(1, len(h3) + 1):
                    ws3.cell(row=row_idx, column=col).fill = fill
            if co.email:
                written_emails.add(co.email.lower())
            row_idx += 1

        for em in c.emails:
            if em.address in written_emails:
                continue
            fill = ALT_FILL if row_idx % 2 == 0 else None
            row_data = [
                c.name, c.company_type or "", c.country or "", c.city or "",
                c.website or "", vtypes,
                "", "", em.address, em.email_type or "general",
                c.source_name or "",
            ]
            ws3.append(row_data)
            if fill:
                for col in range(1, len(h3) + 1):
                    ws3.cell(row=row_idx, column=col).fill = fill
            written_emails.add(em.address)
            row_idx += 1

    autofit(ws3)

    # ---------------------------------------------------------------
    # Sheet 4: By Country pivot
    # ---------------------------------------------------------------
    ws4 = wb.create_sheet("By Country")
    country_data = (
        session.query(Company.country, func.count(Company.id))
        .filter(Company.is_active == True)
        .group_by(Company.country)
        .order_by(func.count(Company.id).desc())
        .all()
    )
    ws4.append(["Country", "Company Count"])
    style_header_row(ws4, 1, 2)
    for country, cnt in country_data:
        ws4.append([country or "Unknown", cnt])
    autofit(ws4)

    wb.save(str(path))
    logger.info("Excel export saved to %s", path)
    return path
